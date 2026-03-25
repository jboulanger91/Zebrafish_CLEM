"""Helper utilities for creating and saving matplotlib figures."""

import os
import pathlib
import sys

import matplotlib.patches as patches
import matplotlib.pyplot as plt
import navis
import numpy as np
import pandas as pd
import seaborn as sns


class Figure:
    """Publication-quality multi-panel figure built on matplotlib.

    Attributes
    ----------
        fig: The underlying matplotlib Figure.
        figure_dict: Dict of resolved plotting parameters.
        figure_data_list: Accumulated data for optional Excel export.
    """

    def __init__(self, **opts_dict):
        """Constructor for a Plotter object.

        Parameters
        ----------
            **opts_dict (keyword arguments): Additional options for the plotter.
                The following options are available:

                - aspect: None or float, optional
                    Aspect ratio of the plot. Default is None.

                - figure_title: str, optional
                    Title of the entire figure. Default is None.

                - plot_title: str, optional
                    Title of the specific plot. Default is None.

                - lc: str, optional
                    Standard color of lines. Default is 'black'.

                - lw: float, optional
                    Standard width of lines. Default is 0.75.

                - line_dashes: tuple or None, optional
                    Line style for plots (e.g., (2, 2) for dashed lines). Default is None.

                - pt: str, optional
                    Standard marker for scatter plots. Default is 'o'.

                - pc: str, optional
                    Standard marker color for scatter plots. Default is 'black'.

                - ps: float, optional
                    Standard marker size for scatter plots. Default is 2.

                - fc: str, optional
                    Fill color of polygons. Default is 'darkgray'.

                - elw: float, optional
                    Standard edge line width for scatter plots and error bars. Default is 0.5.

                - ec: str, optional
                    Standard edge color of markers in scatter plots. Default is 'darkgray'.

                - alpha: float, optional
                    Transparency of elements in the plot. Default is 1.

                - errorbar_area: bool, optional
                    If True, shaded error bars are shown in line plots. Default is True.

                - textcolor: str, optional
                    Color of all text and labels. Default is 'black'.

                - fig_width: float, optional
                    Width of the entire figure in centimeters. Default is 21.00.

                - fig_height: float, optional
                    Height of the entire figure in centimeters. Default is 29.70.

                - helper_lines_dashes: tuple, optional
                    Length of dashes for horizontal and vertical helper lines. Default is (2, 2).

                - helper_lines_lw: float, optional
                    Line width of the helper lines. Default is 0.25 (half of axes_lw).

                - helper_lines_lc: str, optional
                    Color of horizontal and vertical helper lines,
                    and polar grids. Default is 'darkgray'.

                - xl_distance: float, optional
                    Distance (in cm, unless xyzl_3d=True) of the
                    xlabel from the axes. Default is 0.75.

                - yl_distance: float, optional
                    Distance (in cm, unless xyzl_3d=True) of the
                    ylabel from the axes. Default is 0.75.

                - zl_distance: float, optional
                    Distance (in cm, unless xyzl_3d=True) of the
                    zlabel from the axes. Default is 0.75.

                - xyzl_3d: bool
                    If True xl, yl, and zl distances are based on the 3d rotation. Default is False.

                - fontsize: int, optional
                    Font size of all text labels. Default is 6.

                - fontsize_figure_title: int, optional
                    Font size of the figure title on the top of the page. Default is 12.

                - fontsize_plot_label: int, optional
                    Font size of the bold plot labels. Default is 8.

                - dpi: int, optional
                    Dots per inch for rendering the figure. Default is 600.

                - rasterized: bool, optional
                    If True, display plots as pixel rasters for
                    improved rendering and loading of huge data
                    clouds. Default is False.

                - axes_lw: float, optional
                    Line width of all plot borders. Helper lines
                    will be half of this value. Default is 0.5.

                - facecolor: str, optional
                    Background color of the whole plot. Default is 'none'.

                - fontname: str, optional
                    Font name to use for text. Default is 'Arial'.

                - fontfamily: str, optional
                    Font family to use for text. Default is 'sans-serif'.

                - xmin, xmax, ymin, ymax, 3dzmin, 3dzmax: float or None, optional
                    Limits for the x, y, and z axes. Default is None (auto-scaling).

                - vmin, vmax: float or None, optional
                    Limits for colorscaling in scatter3D. Default is None (auto-scaling).

                - elev, azim: float or None, optional
                    Sets the viewpoint of a 3D plot. Default is None (auto-view-init).

                - label: str or None, optional
                    Label for the plot, useful for creating a legend. Default is None.

                - xticks, yticks, zticks: list or None, optional
                    Custom tick positions for the x, y, and z axes.
                    Default is None (auto-generated ticks).

                - xl, yl, zl: str or None, optional
                    Labels for the x, y, and z axes. Default is None.

                - bl: str or None, optional
                    Label for the color bar, useful when using color maps. Default is None.

                - vertical_bar_width: float, optional
                    Width of vertical bars. Default is 0.8.

                - vertical_bar_bottom: float/array, optional
                    y starting value of vertical bars. Default is 0

                - horizontal_bar_height: float, optional
                    Height of horizontal bars. Default is 0.8.

                - horizontal_bar_left: float, optional
                    x-offset of horizontal bars. Default is 0

                - xticklabels_rotation, yticklabels_rotation,
                  zticklabels_rotation: float, optional
                    Rotation angle (in degrees) for tick labels on
                    the x, y, and z axes. Default is 0.

                - zticklabels_ha: string, optional.
                    Horizontal alignment of the tick labels on the z axes. Default is 'left'.

                - vspans, hspans: list or None, optional
                    List of vertical and horizontal spans
                    (rectangles) to add to the plot. Default is None.

                - hlines, vlines: list or None, optional
                    List of horizontal and vertical lines to add to the plot. Default is None.

                - helper_lines_alpha: float, optional
                    Transparency of horizontal and vertical helper
                    lines, and polar grids. Default is 1.0.

                - xlog, ylog, zlog: bool, optional
                    If True, use logarithmic scale for the x, y, and z axes. Default is False.

                - axis_off: bool, optional
                    If True, turn off axis lines and labels. Default is False.

                - plot_label: str or None, optional
                    Label for the specific plot. Default is None.

                - xpos, ypos: float or None, optional
                    X and Y positions for annotations. Default is None.

                - plot_height, plot_width: float or None, optional
                    Height and width of the plot. Default is None (auto-scaling).

                - yerr, xerr, zerr: array-like or None, optional
                    Error bars for the y, x and z axes. Default is None.

                - yerr_pos, yerr_neg: array-like or None, optional
                    Error fill for unequal y error (e.g. quantile range). Default is None.

                - eafc, eaec: string or None, optional
                    Error fillarea face color. Default is None > copies color from 'lc'

                - eaec: string, optional
                    Error fillarea edge color. Default is 'none'

                - eaalpha: float, optional
                    Error fillarea alpha value. Default is 0.2.

                - ealw: float, optional
                    Error fillarea linewidth. Default is 0.

                - legend_xpos, legend_ypos: float or None, optional
                    X and Y positions for the legend. Default is None (let Matplotlib decide).

                - colormap: str, optional
                    Name of the colormap to use for color mappings. Default is 'inferno'.

                - norm_colormap: object or None, optional
                    Normalization object for the colormap. Default is None.

                - show_colormap: bool or None, optional
                    If True, display the color map. Default is None (do not display the color map).

                - image_interpolation: str, optional
                    Interpolation method for displaying images. Default is 'bilinear'.

                - image_origin: str, optional
                    Origin point for displaying images. Default is 'lower'.

                - navis_view: tuple, optional
                    Viewpoint of 3d neuron in 2d. Default is ('x', '-y').

                - navis_alpha: float, optional
                    Alpha value of the brainregions shown in navis plots. Default is 0.2.

                - navis_alpha: str, optional
                    Color of the brainregions shown in navis plots. Default is lightgray.

                - textlabel_ha": str, optional
                    when using function draw_text, horizontal alignment. Default is  "center"

                - textlabel_ma": str, optional
                    when using function draw_text, multi alignment. Default is "center"

                - "textlabel_va": str, optional
                    when using function draw_text, vertical alignment. Default is  "center"

                - "textlabel_rotation": float, optional
                    when using function draw_text, rotation of text, default is 0


        Returns
        -------
            None
        """
        self.figure_dict = {
                 'aspect':None,
                'dash_capstyle':"round",
                 "figure_title": None,  # Title of plot
                 "plot_title": None, # title of the plot
                 "sheet_title": None,  # excel sheet name
                 "ignore_sheet_data": False, # specific plots can not be saved to sheets
                 "lc": 'black',  # Standard color of lines
                 "lw": 0.75,  # Standard width of lines
                 "line_dashes": None,  # Solid if None, e.g. (2,2)
                 "pt": 'o',  # Standard marker for scatter plots
                 "pc": 'black', # Standard marker color
                 "ps": 2,  # Standard marker size for scatter plots,
                 "fc": 'darkgray', # Fillcolor of polygons
                 "elw":0.5,  # Standard edge line width of scatter plots and errorbars
                 "ec": 'darkgray',  # Standard edge color of makers in scatter or quiver plots
                 "alpha": 1,  # Transparency
                 "arrow_scale": None, # quiver plot-specific
                 "arrow_headwidth": 3, # quiver plot-specific
                 "arrow_headlength": 5, # quiver plot-specific
                 "errorbar_area": True,  # Shaded error bar in line plots
                 "textcolor":'black',  # Color of all texts and labnels
                 "fig_width": 21.00,  # Width of the entire figure in cm (21.59 for US-letter)
                 "fig_height": 29.70,  # Height of the entire figure in cm (27.94 for US-letter)
                 "helper_lines_dashes": (2, 2),  # dash lengths
                 "helper_lines_lw": 0.25,  # half of axes_lw
                 "helper_lines_lc": "darkgray",  # helper line color
                 "xl_distance": 0.75,  # xlabel distance (cm)
                 "yl_distance": 0.75,  # ylabel distance (cm)
                 "zl_distance": 0.75,  # zlabel distance (cm)
                 'xyzl_3d': False,  # 3d-based label distances
                 "fontsize": 6,  # Font size of all text labels
                 "fontsize_figure_title": 12,  # figure title size
                 "fontsize_plot_label": 8,  # Font size of the bold plot labels
                 "dpi": 600,
                 "rasterized": False,  # pixel rasters for large data
                 "axes_lw": 0.5,  # border linewidth
                 "facecolor": 'none', # The background color of the whole plot
                 "fontname": 'Arial',
                 "fontfamily": 'sans-serif',
                 "xmin": None,
                 "xmax": None,
                 "ymin": None,
                 "ymax": None,
                 "3dzmin": None,
                 "3dzmax": None,
                 "zmin": None,
                 "zmax": None,
                 "vmin": None,
                 "vmax": None,
                 "elev": None,
                 "azim": None,
                 "label": None,
                 "sheet_label": None,  # excel sheet label
                 "xticks": None,
                 "yticks": None,
                 "zticks": None,
                 "xl": None,
                 "yl": None,
                 "zl": None,
                 "bl": None,
                 "vertical_bar_width": 0.8,
                 "vertical_bar_bottom": 0,
                 "horizontal_bar_height": 0.8,
                 "horizontal_bar_left": 0,
                 "xticklabels": None,
                 "xticklabels_rotation": 0,
                 "yticklabels": None,
                 "yticklabels_rotation": 0,
                 "zticklabels": None,
                 "zticklabels_rotation": 0,
                 "zticklabels_ha": 'left',
                 "vspans": None,
                 "hspans": None,
                 "hlines": None,
                 "helper_lines_alpha": 1.0,
                 "vlines": None,
                 "xlog": False,
                 "ylog": False,
                 "zlog": False,
                 "axis_off": False,
                 "plot_label": None,
                 "xpos": None,
                "ypos": None,
                'plot_height': None,
                'plot_width': None,
                'yerr': None,
                'yerr_pos': None,
                'yerr_neg': None,
                'xerr': None,
                'zerr': None,
                'eafc': None,
                'eaalpha': 0.2,
                'ealw': 0,
                'eaec': 'none',
                 "legend_xpos": None,
                 "legend_ypos": None,
                 "colormap": "inferno",
                 "norm_colormap": None,
                 "show_colormap": None,
                 "image_interpolation": "bilinear",
                 "image_origin": "lower",
                 "navis_view": ("x", "-y"),
                 "navis_alpha": 0.2,
                 "navis_color": 'lightgray',
                 "textlabel_ha": "center",
                "textlabel_ma": "center",
                "textlabel_va": "center",
                "textlabel_rotation": 0}

        for key in opts_dict:
            if key not in self.figure_dict:
                raise ValueError(key, "is not a valid plotting parameter.")
            self.figure_dict[key] = opts_dict[key]

        plt.rcParams['font.sans-serif'] = self.figure_dict["fontname"]
        plt.rcParams['font.family'] = self.figure_dict["fontfamily"]
        plt.rcParams['font.size'] = self.figure_dict["fontsize"]
        plt.rcParams['figure.dpi'] = self.figure_dict["dpi"]
        plt.rcParams['pdf.fonttype'] = 42  # Always embedd fonts in the pdf
        plt.rcParams['ps.fonttype'] = 42

        # Figure size must be in inches
        self.fig = plt.figure(
            num=None,
            facecolor=self.figure_dict["facecolor"],
            edgecolor='none',
            dpi=self.figure_dict["dpi"],
        )

        # Somhow on MAC, one needs to do this
        self.fig.set_size_inches(
            self.figure_dict["fig_width"] / 2.54,
            self.figure_dict["fig_height"] / 2.54,
        )

        if self.figure_dict["figure_title"] is not None:
            plt.figtext(
                0.5, 0.95,
                self.figure_dict["figure_title"],
                ha='center', va='center',
                fontsize=self.figure_dict["fontsize_figure_title"],
                color=self.figure_dict["textcolor"],
            )

        # Keeps track of all figure data
        self.figure_data_list = []

    def create_plot(self, **opts_dict):
        """Create a 2D cartesian subplot panel."""
        return Plot(self, opts_dict)

    def create_polar_plot(self, **opts_dict):
        """Create a polar subplot panel."""
        return PolarPlot(self, opts_dict)

    def create_plot3D(self, **opts_dict):
        """Create a 3D subplot panel."""
        return Plot3D(self, opts_dict)

    def show(self):
        """Display the figure in the default matplotlib backend."""
        self.fig.show()

    def save(self, path, tight=False, open_file=False, save_data_as_excel=False):
        """Save the figure to disk and optionally export data as Excel.

        Args:
            path: Output file path (e.g. PDF, PNG).
            tight: If True, use tight bounding box.
            open_file: If True, open the saved file in the system viewer.
            save_data_as_excel: If True, write plot data to an .xlsx sidecar file.
        """
        if tight:
            self.fig.savefig(
                path, bbox_inches='tight', pad_inches=0.1,
                facecolor=self.figure_dict["facecolor"],
                edgecolor='none', transparent=True,
            )
        else:
            self.fig.savefig(
                path,
                facecolor=self.figure_dict["facecolor"],
                edgecolor='none', transparent=True,
            )

        if save_data_as_excel:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter

            # Create a new workbook
            wb = Workbook()

            # Remove the default sheet if you want
            default_sheet = wb.active
            wb.remove(default_sheet)

            for plot_data in self.figure_data_list:
                sheet = wb.create_sheet(plot_data["title"])

                j = 1
                for label_i in range(len(plot_data["label"])):

                    sheet[f"{get_column_letter(j)}1"] = plot_data["label"][label_i]

                    sheet[f"{get_column_letter(j)}2"] = plot_data["xl"]
                    for i, x in enumerate(plot_data["x"][label_i]):
                        sheet[f"{get_column_letter(j)}{3+i}"] = x

                    j += 1

                    sheet[f"{get_column_letter(j)}2"] = plot_data["yl"]
                    for i, y in enumerate(plot_data["y"][label_i]):
                        sheet[f"{get_column_letter(j)}{3+i}"] = y

                    j += 1

                    if len(plot_data["z"]) > 0:
                        sheet[f"{get_column_letter(j)}2"] = plot_data["zl"]

                        for i, z in enumerate(plot_data["z"][label_i]):
                            sheet[f"{get_column_letter(j)}{3+i}"] = z

                        j += 1

                    j += 1

            wb.save(pathlib.Path(path).with_suffix(".xlsx"))

        plt.close('all')

        if open_file:
            if sys.platform.startswith('darwin'):
                os.system(f"open '{path}'")
            else:
                os.startfile(path)



    def add_text(self, x, y, text, rotation=0, ha='center'):
        """Place text on the figure at absolute cm coordinates."""
        plt.figtext(
            x / self.figure_dict["fig_width"],
            y / self.figure_dict["fig_height"],
            text, ha=ha, ma=ha, va='center',
            color=self.figure_dict["textcolor"],
            rotation=rotation,
        )


class SuperPlot:
    """Base class for all plot panel types providing shared drawing methods."""

    def __init__(self, figure, opts_dict):
        """Initialize plot panel with resolved options from the parent Figure."""
        self.figure = figure
        self.plot_dict = self.figure.figure_dict.copy()

        for key in opts_dict:
            if key not in self.plot_dict:
                raise ValueError(key, "is not a valid plotting parameter.")
            self.plot_dict[key] = opts_dict[key]

        if self.plot_dict["legend_xpos"] is None:
            self.plot_dict["legend_xpos"] = self.plot_dict["xpos"] + self.plot_dict["plot_width"]
        if self.plot_dict["legend_ypos"] is None:
            self.plot_dict["legend_ypos"] = self.plot_dict["ypos"] + self.plot_dict["plot_height"]

        self.current_zorder = 0

        # Each panel adds an index, because panels do not have name ID
        # Plot data is a dict, it will later be filled by the line drawing etc
        self.plot_data = {}

        if self.plot_dict["ignore_sheet_data"] is False:
            if self.plot_dict["sheet_title"] is not None:
                self.plot_data["title"] = self.plot_dict["sheet_title"]
            else:
                self.plot_data["title"] = self.plot_dict["plot_title"]

            self.plot_data["xl"] = self.plot_dict["xl"]
            self.plot_data["yl"] = self.plot_dict["yl"]
            self.plot_data["zl"] = self.plot_dict["zl"]

            self.plot_data["x"] = []
            self.plot_data["y"] = []
            self.plot_data["z"] = []
            self.plot_data["label"] = []

            self.figure.figure_data_list.append(self.plot_data)


    def set_axes_properties(self):
        """Apply axis limits, labels, scales, and visibility from plot_dict."""
        if self.plot_dict["xmin"] is not None:
            self.ax.set_xlim([self.plot_dict["xmin"], self.ax.get_xlim()[1]])
        if self.plot_dict["ymin"] is not None:
            self.ax.set_ylim([self.plot_dict["ymin"], self.ax.get_ylim()[1]])
        if self.plot_dict["3dzmin"] is not None:
            self.ax.set_zlim([self.plot_dict["3dzmin"], self.ax.get_zlim()[1]])

        if self.plot_dict["xmax"] is not None:
            self.ax.set_xlim([self.ax.get_xlim()[0], self.plot_dict["xmax"]])
        if self.plot_dict["ymax"] is not None:
            self.ax.set_ylim([self.ax.get_ylim()[0], self.plot_dict["ymax"]])
        if self.plot_dict["3dzmax"] is not None:
            self.ax.set_zlim([self.ax.get_zlim()[0], self.plot_dict["3dzmax"]])


        if self.plot_dict["xl"] is not None:
            self.ax.set_xlabel(
                self.plot_dict["xl"],
                horizontalalignment='center',
                verticalalignment='center',
                color=self.plot_dict["textcolor"],
            )

            if self.plot_dict['xyzl_3d']:
                self.ax.xaxis.labelpad = self.plot_dict["xl_distance"]
            else:
                x_coord = (
                    (self.plot_dict["xpos"]
                     + 0.5 * self.plot_dict["plot_width"])
                    / self.plot_dict["fig_width"]
                )
                y_coord = (
                    (self.plot_dict["ypos"]
                     - self.plot_dict["xl_distance"])
                    / self.plot_dict["fig_height"]
                )

                self.ax.xaxis.set_label_coords(
                    x_coord, y_coord,
                    self.figure.fig.transFigure,
                )

        if self.plot_dict["yl"] is not None:
            self.ax.set_ylabel(
                self.plot_dict["yl"],
                verticalalignment='center',
                horizontalalignment='center',
                color=self.plot_dict["textcolor"],
            )

            if self.plot_dict['xyzl_3d']:
                self.ax.yaxis.labelpad = self.plot_dict["yl_distance"]
            else:
                x_coord = (
                    (self.plot_dict["xpos"]
                     - self.plot_dict["yl_distance"])
                    / self.plot_dict["fig_width"]
                )
                y_coord = (
                    (self.plot_dict["ypos"]
                     + 0.5 * self.plot_dict["plot_height"])
                    / self.plot_dict["fig_height"]
                )

                self.ax.yaxis.set_label_coords(
                    x_coord, y_coord,
                    self.figure.fig.transFigure,
                )

        if (
            self.plot_dict["zl"] is not None
            and self.plot_dict["show_colormap"] is None
        ):
            self.ax.set_zlabel(
                self.plot_dict["zl"],
                verticalalignment='center',
                horizontalalignment='center',
                color=self.plot_dict["textcolor"],
            )

            if self.plot_dict['xyzl_3d']:
                self.ax.zaxis.labelpad = self.plot_dict["zl_distance"]

        if self.plot_dict["xlog"] is True:
            self.ax.set_xscale("log")

        if self.plot_dict["ylog"] is True:
            self.ax.set_yscale("log")

        if self.plot_dict["zlog"] is True:
            self.ax.set_zscale("log")

        if self.plot_dict["axis_off"] is True:
            self.ax.set_axis_off()

    def draw_horizontal_significance_label(self, x0, x1, y, label):
        """Draw a horizontal bracket with a centered text label above it."""
        self.current_zorder += 1
        self.ax.plot(
            [x0, x1], [y, y],
            color=self.plot_dict["textcolor"],
            lw=self.plot_dict["axes_lw"],
            alpha=1.0, solid_capstyle="round",
            rasterized=self.plot_dict["rasterized"],
            zorder=self.current_zorder,
        )

        self.ax.text(
            (x0 + x1) / 2, y, label,
            ha='center', ma='center', va='bottom',
            color=self.plot_dict["textcolor"],
            rotation=0, zorder=self.current_zorder,
        )

    def draw_vertical_significance_label(self, y0, y1, x, label):
        """Draw a vertical bracket with a text label to the right."""
        self.current_zorder += 1
        self.ax.plot(
            [x, x], [y0, y1],
            color=self.plot_dict["textcolor"],
            lw=self.plot_dict["axes_lw"],
            alpha=1.0, solid_capstyle="round",
            rasterized=self.plot_dict["rasterized"],
            zorder=self.current_zorder,
        )

        self.ax.text(
            x, (y0 + y1) / 2, " " + label,
            ha='left', ma='left', va='center',
            color=self.plot_dict["textcolor"],
            rotation=0., zorder=self.current_zorder,
        )

    def draw_rectangle(self, x, y, width, height, fill, **opts_dict):
        """Draw a rectangle patch on the axes."""
        artist_dict = self.plot_dict.copy()

        for key in opts_dict:
            if key not in artist_dict:
                raise ValueError(key, "is not a valid plotting parameter.")

            artist_dict[key] = opts_dict[key]

        self.current_zorder += 1

        if artist_dict["line_dashes"] is not None:
            rect = patches.Rectangle((x, y), width, height, fill=fill,
                                     linewidth=artist_dict["lw"],
                                     edgecolor=artist_dict["lc"],
                                     facecolor=artist_dict["fc"],
                                     linestyle=(0, artist_dict["line_dashes"]),
                                     capstyle="round",
                                     alpha=artist_dict["alpha"],
                                     rasterized=artist_dict["rasterized"],
                                     zorder=self.current_zorder )

        else:
            rect = patches.Rectangle((x, y), width, height, fill=fill,
                                     linewidth=artist_dict["lw"],
                                     edgecolor=artist_dict["lc"],
                                     facecolor=artist_dict["fc"],
                                     alpha=artist_dict["alpha"],
                                     rasterized=artist_dict["rasterized"],
                                     zorder=self.current_zorder)

        self.ax.add_artist(rect)

    def draw_circle(self, x, y, r, fill, **opts_dict):
        """Draw a circle patch on the axes."""
        artist_dict = self.plot_dict.copy()

        for key in opts_dict:
            if key not in artist_dict:
                raise ValueError(key, "is not a valid plotting parameter.")

            artist_dict[key] = opts_dict[key]
        self.current_zorder += 1
        if artist_dict["line_dashes"] is not None:
            circle = plt.Circle(
                (x, y), r, fill=fill,
                edgecolor=artist_dict["lc"],
                fc=artist_dict["fc"],
                linestyle=(0, artist_dict["line_dashes"]),
                capstyle="round",
                alpha=artist_dict["alpha"],
                lw=artist_dict["lw"],
                rasterized=artist_dict["rasterized"],
                zorder=self.current_zorder,
            )
        else:
            circle = plt.Circle(
                (x, y), r, fill=fill,
                edgecolor=artist_dict["lc"],
                fc=artist_dict["fc"],
                alpha=artist_dict["alpha"],
                lw=artist_dict["lw"],
                rasterized=artist_dict["rasterized"],
                zorder=self.current_zorder,
            )

        self.ax.add_artist(circle)

    def draw_horizontal_statistics_line(self, x0, x1, y, p, **opts_dict):
        """Draw a horizontal significance bracket with star-notation for a p-value."""
        artist_dict = self.plot_dict.copy()

        for key in opts_dict:
            if key not in artist_dict:
                raise ValueError(key, "is not a valid plotting parameter.")
            artist_dict[key] = opts_dict[key]

        # This is just a helper line, dont store the datasheet info for this
        artist_dict["ignore_sheet_data"] = True
        self.draw_line(x=[x0, x1], y=[y, y], **artist_dict)

        if p < 0.001:
            stats_str = "***"
        elif p < 0.01:
            stats_str = "**"
        elif p < 0.05:
            stats_str = "*"
        else:
            stats_str = "ns"

        # Always make it this bottom aligned here, so the stars look fine
        artist_dict['textlabel_va'] = 'bottom'

        self.draw_text(x=(x0+x1) / 2, y=y, text=stats_str, **artist_dict)

    def draw_text(self, x, y, text, **opts_dict):
        """Draw a text label at the given data coordinates."""
        artist_dict = self.plot_dict.copy()

        self.current_zorder += 1
        for key in opts_dict:
            if key not in artist_dict:
                raise ValueError(key, "is not a valid plotting parameter.")
            artist_dict[key] = opts_dict[key]

        self.ax.text(x, y, text,
                     ha=artist_dict["textlabel_ha"],
                     ma=artist_dict["textlabel_ma"],
                     va=artist_dict["textlabel_va"],
                     color=artist_dict["textcolor"],
                     rotation=artist_dict["textlabel_rotation"],
                     zorder=self.current_zorder)

    def draw_line(self, x, y, **opts_dict):
        """Draw a line plot with optional error bands or error bars."""
        artist_dict = self.plot_dict.copy()

        if isinstance(x, (int, float)):
            x = [x]
        if isinstance(y, (int, float)):
            y = [y]

        for key in opts_dict:
            if key not in artist_dict:
                raise ValueError(key, "is not a valid plotting parameter.")
            artist_dict[key] = opts_dict[key]

        if artist_dict["ignore_sheet_data"] is False:
            if artist_dict["sheet_label"] is not None:
                self.plot_data["label"].append(artist_dict["sheet_label"])
            else:
                self.plot_data["label"].append(artist_dict["label"])

            self.plot_data["x"].append(x)
            self.plot_data["y"].append(y)

        if  artist_dict['yerr'] is not None:
            yerr = np.array(artist_dict["yerr"])
            if artist_dict["eafc"] is None:
                artist_dict["eafc"] = artist_dict["lc"]
            self.current_zorder += 1
            if artist_dict["errorbar_area"] is False:
                self.ax.errorbar(
                    x, y, yerr=yerr,
                    elinewidth=artist_dict["elw"],
                    ecolor=artist_dict["lc"],
                    fmt='none',
                    capsize=1.5 * artist_dict["elw"],
                    mew=artist_dict["elw"],
                    solid_capstyle='round',
                    solid_joinstyle='round',
                    rasterized=artist_dict["rasterized"],
                    zorder=self.current_zorder,
                )
            else:
                self.ax.fill_between(
                    x, y - yerr, y + yerr,
                    lw=artist_dict["ealw"],
                    edgecolor=artist_dict["eaec"],
                    facecolor=artist_dict["eafc"],
                    alpha=artist_dict["eaalpha"],
                    rasterized=artist_dict["rasterized"],
                    zorder=self.current_zorder,
                )

        if (artist_dict['yerr_pos'] is not None
                and artist_dict['yerr_neg'] is not None):
            yerr_pos = np.array(artist_dict["yerr_pos"])
            yerr_neg = np.array(artist_dict["yerr_neg"])
            if artist_dict["eafc"] is None:
                artist_dict["eafc"] = artist_dict["lc"]
            self.current_zorder += 1

            if artist_dict["errorbar_area"] is False:
                self.ax.errorbar(
                    x, y,
                    yerr=[yerr_neg, yerr_pos],
                    elinewidth=artist_dict["elw"],
                    ecolor=artist_dict["lc"],
                    fmt='none',
                    capsize=1.5 * artist_dict["elw"],
                    mew=artist_dict["elw"],
                    solid_capstyle='round',
                    solid_joinstyle='round',
                    rasterized=artist_dict["rasterized"],
                    zorder=self.current_zorder,
                )
            else:
                self.ax.fill_between(
                    x, y - yerr_neg, y + yerr_pos,
                    lw=artist_dict["ealw"],
                    edgecolor=artist_dict["eaec"],
                    facecolor=artist_dict["eafc"],
                    alpha=artist_dict["eaalpha"],
                    rasterized=artist_dict["rasterized"],
                    zorder=self.current_zorder,
                )

        if artist_dict["xerr"] is not None:
            xerr = np.array(artist_dict["xerr"])
            if artist_dict["eafc"] is None:
                artist_dict["eafc"] = artist_dict["lc"]
            self.current_zorder += 1
            if artist_dict["errorbar_area"] is False:
                self.ax.errorbar(
                    x, y, xerr=xerr,
                    elinewidth=artist_dict["elw"],
                    ecolor=artist_dict["lc"],
                    fmt='none',
                    capsize=1.5 * artist_dict["elw"],
                    mew=artist_dict["elw"],
                    solid_capstyle='round',
                    solid_joinstyle='round',
                    rasterized=artist_dict["rasterized"],
                    zorder=self.current_zorder,
                )
            else:
                self.ax.fill_betweenx(
                    y, x + xerr, x - xerr,
                    lw=artist_dict["ealw"],
                    edgecolor=artist_dict["eaec"],
                    facecolor=artist_dict["eafc"],
                    alpha=artist_dict["eaalpha"],
                    rasterized=artist_dict["rasterized"],
                    zorder=self.current_zorder,
                )

        self.current_zorder += 1

        if artist_dict["line_dashes"] is not None:
            self.ax.plot(
                x, y,
                color=artist_dict["lc"],
                lw=artist_dict["lw"],
                alpha=artist_dict["alpha"],
                dashes=artist_dict["line_dashes"],
                dash_capstyle=artist_dict["dash_capstyle"],
                label=artist_dict["label"],
                rasterized=artist_dict["rasterized"],
                zorder=self.current_zorder,
            )
        else:
            self.ax.plot(
                x, y,
                color=artist_dict["lc"],
                lw=artist_dict["lw"],
                alpha=artist_dict["alpha"],
                solid_capstyle="round",
                label=artist_dict["label"],
                rasterized=artist_dict["rasterized"],
                zorder=self.current_zorder,
            )

        if artist_dict["label"] is not None:
            leg = self.ax.legend(
                frameon=False, loc='upper left',
                bbox_to_anchor=(
                    artist_dict["legend_xpos"]
                    / artist_dict["fig_width"],
                    artist_dict["legend_ypos"]
                    / artist_dict["fig_height"],
                ),
                bbox_transform=self.figure.fig.transFigure,
            )

            for text in leg.get_texts():
                plt.setp(text, color=artist_dict["textcolor"])

    def draw_polygon(self, x, y, **opts_dict):
        """Draw a filled polygon from vertex coordinates."""
        artist_dict = self.plot_dict.copy()

        if isinstance(x, (int, float)):
            x = [x]
        if isinstance(y, (int, float)):
            y = [y]

        for key in opts_dict:
            if key not in artist_dict:
                raise ValueError(key, "is not a valid plotting parameter.")
            artist_dict[key] = opts_dict[key]

        if artist_dict["ignore_sheet_data"] is False:
            if artist_dict["sheet_label"] is not None:
                self.plot_data["label"].append(artist_dict["sheet_label"])
            else:
                self.plot_data["label"].append(artist_dict["label"])

            self.plot_data["x"].append(x)
            self.plot_data["y"].append(y)

        self.current_zorder += 1

        if artist_dict["line_dashes"] is not None:
            self.ax.fill(
                x, y, fc=artist_dict["fc"],
                alpha=artist_dict["alpha"],
                ec=None, lw=None,
                label=artist_dict["label"],
                rasterized=artist_dict["rasterized"],
                zorder=self.current_zorder,
            )
            self.ax.plot(
                x, y, color=artist_dict["lc"],
                lw=artist_dict["lw"],
                dashes=artist_dict["line_dashes"],
                dash_capstyle=artist_dict["dash_capstyle"],
                rasterized=artist_dict["rasterized"],
                zorder=self.current_zorder,
            )

        else:
            self.ax.fill(
                x, y, fc=artist_dict["fc"],
                alpha=artist_dict["alpha"],
                ec=None, lw=None,
                label=artist_dict["label"],
                rasterized=artist_dict["rasterized"],
                zorder=self.current_zorder,
            )
            self.ax.plot(
                x, y, color=artist_dict["lc"],
                lw=artist_dict["lw"],
                solid_capstyle="round",
                rasterized=artist_dict["rasterized"],
                zorder=self.current_zorder,
            )

        if artist_dict["label"] is not None:
            leg = self.ax.legend(
                frameon=False, loc='upper left',
                bbox_to_anchor=(
                    artist_dict["legend_xpos"]
                    / artist_dict["fig_width"],
                    artist_dict["legend_ypos"]
                    / artist_dict["fig_height"],
                ),
                bbox_transform=self.figure.fig.transFigure,
            )

            for text in leg.get_texts():
                plt.setp(text, color=artist_dict["textcolor"])

    def draw_scatter(self, x, y, **opts_dict):
        """Draw a 2D scatter plot with optional error bars and colormap support."""
        artist_dict = self.plot_dict.copy()

        if isinstance(x, (int, float)):
            x = [x]
        if isinstance(y, (int, float)):
            y = [y]

        for key in opts_dict:
            if key not in artist_dict:
                raise ValueError(key, "is not a valid plotting parameter.")
            artist_dict[key] = opts_dict[key]

        if artist_dict["ignore_sheet_data"] is False:
            if artist_dict["sheet_label"] is not None:
                self.plot_data["label"].append(artist_dict["sheet_label"])
            else:
                self.plot_data["label"].append(artist_dict["label"])

            self.plot_data["x"].append(x)
            self.plot_data["y"].append(y)

        if artist_dict["yerr"] is not None:
            self.current_zorder += 1
            self.ax.errorbar(
                x, y,
                yerr=artist_dict["yerr"],
                elinewidth=artist_dict["elw"],
                ecolor=artist_dict["ec"],
                fmt='none',
                capsize=1.5 * artist_dict["elw"],
                mew=artist_dict["elw"],
                solid_capstyle='round',
                solid_joinstyle='round',
                zorder=self.current_zorder,
            )

        if (artist_dict['yerr_pos'] is not None
                and artist_dict['yerr_neg'] is not None):
            yerr_pos = np.array(artist_dict["yerr_pos"])
            yerr_neg = np.array(artist_dict["yerr_neg"])
            self.current_zorder += 1
            self.ax.errorbar(
                x, y,
                yerr=[yerr_neg, yerr_pos],
                elinewidth=artist_dict["elw"],
                ecolor=artist_dict["ec"],
                fmt='none',
                capsize=1.5 * artist_dict["elw"],
                mew=artist_dict["elw"],
                solid_capstyle='round',
                solid_joinstyle='round',
                zorder=self.current_zorder,
            )

        if artist_dict["xerr"] is not None:
            self.current_zorder += 1
            self.ax.errorbar(
                x, y,
                xerr=artist_dict["xerr"],
                elinewidth=artist_dict["elw"],
                ecolor=artist_dict["ec"],
                fmt='none',
                capsize=1.5 * artist_dict["elw"],
                mew=artist_dict["elw"],
                solid_capstyle='round',
                solid_joinstyle='round',
                zorder=self.current_zorder,
            )

        # to get the same units as seaborn, we need to use the square of the marker size
        self.current_zorder += 1

        if isinstance(artist_dict["pc"], (list, np.ndarray)):
            if artist_dict["norm_colormap"] is None:
                self.ax.scatter(
                    x, y, c=artist_dict["pc"],
                    marker=artist_dict["pt"],
                    s=artist_dict["ps"] ** 2,
                    linewidths=artist_dict["elw"],
                    edgecolor=artist_dict["ec"],
                    alpha=artist_dict["alpha"],
                    label=artist_dict["label"],
                    rasterized=artist_dict["rasterized"],
                    cmap=artist_dict["colormap"],
                    norm=None,
                    vmin=artist_dict["zmin"],
                    vmax=artist_dict["zmax"],
                    zorder=self.current_zorder,
                )
            else:
                self.ax.scatter(
                    x, y, c=artist_dict["pc"],
                    marker=artist_dict["pt"],
                    s=artist_dict["ps"] ** 2,
                    linewidths=artist_dict["elw"],
                    edgecolor=artist_dict["ec"],
                    alpha=artist_dict["alpha"],
                    label=artist_dict["label"],
                    rasterized=artist_dict["rasterized"],
                    cmap=artist_dict["colormap"],
                    norm=artist_dict["norm_colormap"],
                    zorder=self.current_zorder,
                )

        else:
            self.ax.scatter(
                x, y, c=artist_dict["pc"],
                marker=artist_dict["pt"],
                s=artist_dict["ps"] ** 2,
                linewidths=artist_dict["elw"],
                edgecolor=artist_dict["ec"],
                alpha=artist_dict["alpha"],
                label=artist_dict["label"],
                rasterized=artist_dict["rasterized"],
                zorder=self.current_zorder,
            )

        if artist_dict["label"] is not None:
            leg = self.ax.legend(
                frameon=False, loc='upper left',
                bbox_to_anchor=(
                    artist_dict["legend_xpos"]
                    / artist_dict["fig_width"],
                    artist_dict["legend_ypos"]
                    / artist_dict["fig_height"],
                ),
                bbox_transform=self.figure.fig.transFigure,
            )

            for text in leg.get_texts():
                plt.setp(text, color=artist_dict["textcolor"])

    def draw_scatter3D(self, x, y, z, **opts_dict):
        """Draw a 3D scatter plot with optional error bars and colormap support."""
        artist_dict = self.plot_dict.copy()

        if isinstance(x, (int, float)):
            x = [x]
        if isinstance(y, (int, float)):
            y = [y]
        if isinstance(z, (int, float)):
            z = [z]

        for key in opts_dict:
            if key not in artist_dict:
                raise ValueError(key, "is not a valid plotting parameter.")
            artist_dict[key] = opts_dict[key]

        if artist_dict["ignore_sheet_data"] is False:
            if artist_dict["sheet_label"] is not None:
                self.plot_data["label"].append(artist_dict["sheet_label"])
            else:
                self.plot_data["label"].append(artist_dict["label"])

            self.plot_data["x"].append(x)
            self.plot_data["y"].append(y)
            self.plot_data["z"].append(z)

        if artist_dict["yerr"] is not None:
            self.current_zorder += 1
            self.ax.errorbar(
                x, y, z,
                yerr=artist_dict["yerr"],
                elinewidth=artist_dict["elw"],
                ecolor=artist_dict["ec"],
                fmt='none',
                capsize=1.5 * artist_dict["elw"],
                mew=artist_dict["elw"],
                solid_capstyle='round',
                solid_joinstyle='round',
                zorder=self.current_zorder,
            )

        if artist_dict["xerr"] is not None:
            self.current_zorder += 1
            self.ax.errorbar(
                x, y, z,
                xerr=artist_dict["xerr"],
                elinewidth=artist_dict["elw"],
                ecolor=artist_dict["ec"],
                fmt='none',
                capsize=1.5 * artist_dict["elw"],
                mew=artist_dict["elw"],
                solid_capstyle='round',
                solid_joinstyle='round',
                zorder=self.current_zorder,
            )

        if artist_dict["zerr"] is not None:
            self.current_zorder += 1
            self.ax.errorbar(
                x, y, z,
                zerr=artist_dict["zerr"],
                elinewidth=artist_dict["elw"],
                ecolor=artist_dict["ec"],
                fmt='none',
                capsize=1.5 * artist_dict["elw"],
                mew=artist_dict["elw"],
                solid_capstyle='round',
                solid_joinstyle='round',
                zorder=self.current_zorder,
            )

        # to get the same units as seaborn, we need to use the square of the marker size
        self.current_zorder += 1

        if isinstance(artist_dict["pc"], (list, np.ndarray)):
            if artist_dict["norm_colormap"] is None:
                self.ax.scatter(
                    x, y, z, c=artist_dict["pc"],
                    marker=artist_dict["pt"],
                    s=artist_dict["ps"] ** 2,
                    linewidths=artist_dict["elw"],
                    edgecolor=artist_dict["ec"],
                    alpha=artist_dict["alpha"],
                    label=artist_dict["label"],
                    rasterized=artist_dict["rasterized"],
                    cmap=artist_dict["colormap"],
                    norm=None,
                    vmin=artist_dict["vmin"],
                    vmax=artist_dict["vmax"],
                    zorder=self.current_zorder,
                )
            else:
                self.ax.scatter(
                    x, y, z, c=artist_dict["pc"],
                    marker=artist_dict["pt"],
                    s=artist_dict["ps"] ** 2,
                    linewidths=artist_dict["elw"],
                    edgecolor=artist_dict["ec"],
                    alpha=artist_dict["alpha"],
                    label=artist_dict["label"],
                    rasterized=artist_dict["rasterized"],
                    cmap=artist_dict["colormap"],
                    norm=artist_dict["norm_colormap"],
                    zorder=self.current_zorder,
                )

        else:
            self.ax.scatter(
                x, y, z, c=artist_dict["pc"],
                marker=artist_dict["pt"],
                s=artist_dict["ps"] ** 2,
                linewidths=artist_dict["elw"],
                edgecolor=artist_dict["ec"],
                alpha=artist_dict["alpha"],
                label=artist_dict["label"],
                rasterized=artist_dict["rasterized"],
                zorder=self.current_zorder,
            )

        if artist_dict["label"] is not None:
            leg = self.ax.legend(
                frameon=False, loc='upper left',
                bbox_to_anchor=(
                    artist_dict["legend_xpos"]
                    / artist_dict["fig_width"],
                    artist_dict["legend_ypos"]
                    / artist_dict["fig_height"],
                ),
                bbox_transform=self.figure.fig.transFigure,
            )

            for text in leg.get_texts():
                plt.setp(text, color=artist_dict["textcolor"])

    def draw_quiver(self, XX, YY, U, V, **opts_dict):
        """Draw a quiver (vector field) plot."""
        artist_dict = self.plot_dict.copy()

        for key in opts_dict:
            if key not in artist_dict:
                raise ValueError(key, "is not a valid plotting parameter.")

            artist_dict[key] = opts_dict[key]

        self.current_zorder += 1

        self.ax.quiver(XX, YY, U, V,
                       scale=artist_dict["arrow_scale"],
                       headwidth=artist_dict["arrow_headwidth"],
                       headlength=artist_dict["arrow_headlength"],
                       color=artist_dict["lc"],
                       rasterized=artist_dict["rasterized"],
                       linewidths=artist_dict["lw"],
                       edgecolor=artist_dict["ec"],
                       zorder=self.current_zorder)

    def draw_vertical_bars(self, x, y, **opts_dict):
        """Draw vertical bar chart with optional error bars and labels."""
        artist_dict = self.plot_dict.copy()

        if isinstance(x, (int, float)):
            x = [x]
        if isinstance(y, (int, float)):
            y = [y]

        for key in opts_dict:
            if key not in artist_dict:
                raise ValueError(key, "is not a valid plotting parameter.")
            artist_dict[key] = opts_dict[key]

        if artist_dict["ignore_sheet_data"] is False:
            if artist_dict["sheet_label"] is not None:
                self.plot_data["label"].append(artist_dict["sheet_label"])
            else:
                self.plot_data["label"].append(artist_dict["label"])

            self.plot_data["x"].append(x)
            self.plot_data["y"].append(y)

        if type(artist_dict["lw"]) is not list:
            artist_dict["lw"] = [artist_dict["lw"]] * len(x)

        if type(artist_dict["lc"]) is not list:
            artist_dict["lc"] = [artist_dict["lc"]] * len(x)

        if type(artist_dict["alpha"]) is not list:
            artist_dict["alpha"] = [artist_dict["alpha"]] * len(x)

        if type(artist_dict['vertical_bar_bottom']) is not list:
            artist_dict['vertical_bar_bottom'] = (
                [artist_dict['vertical_bar_bottom']] * len(x)
            )

        for i in range(len(x)):

            if artist_dict["yerr"] is not None:
                self.current_zorder += 1
                self.ax.errorbar(
                    x[i], y[i],
                    yerr=artist_dict["yerr"][i],
                    elinewidth=artist_dict["elw"],
                    ecolor=artist_dict["lc"][i],
                    fmt='none',
                    rasterized=artist_dict["rasterized"],
                    capsize=artist_dict["elw"] * 1.5,
                    mew=artist_dict["elw"],
                    solid_capstyle='round',
                    solid_joinstyle='round',
                    zorder=self.current_zorder,
                )

            self.current_zorder += 1

            bar_label = (
                artist_dict["label"]
                if i == 0 and artist_dict["label"] is not None
                else None
            )
            self.ax.bar(
                x[i], y[i],
                edgecolor=None, lw=None,
                alpha=artist_dict["alpha"][i],
                facecolor=artist_dict["lc"][i],
                rasterized=artist_dict["rasterized"],
                align='center',
                width=artist_dict["vertical_bar_width"],
                label=bar_label,
                zorder=self.current_zorder,
                bottom=artist_dict["vertical_bar_bottom"][i],
            )

            if (
                artist_dict["bl"] is not None
                and not np.isnan(y[0])
                and not np.isnan(x[0])
            ):
                x_ = (
                    self.ax.xpos / artist_dict["fig_width"]
                    + artist_dict["plot_width"]
                    * ((x[0] - self.ax.xmin)
                       / (self.ax.xmax - self.ax.xmin))
                    / artist_dict["fig_width"]
                )
                yerr_val = artist_dict["yerr"][0]
                y_ = (
                    self.ax.ypos / artist_dict["fig_height"]
                    + artist_dict["plot_heighth"]
                    * ((y[0]
                        + np.sign(y[0] + yerr_val) * yerr_val
                        - self.ax.ymin)
                       / (self.ax.ymax - self.ax.ymin))
                    / artist_dict["fig_height"]
                    + np.sign(y[0] + yerr_val)
                    * 0.2 / artist_dict["fig_height"]
                )

                plt.figtext(
                    x_, y_, artist_dict["bl"],
                    ha='center', va='center',
                    color=artist_dict["textcolor"],
                )

        if artist_dict["label"] is not None:
            leg = self.ax.legend(
                frameon=False, loc='upper left',
                bbox_to_anchor=(
                    artist_dict["legend_xpos"]
                    / artist_dict["fig_width"],
                    artist_dict["legend_ypos"]
                    / artist_dict["fig_height"],
                ),
                bbox_transform=self.figure.fig.transFigure,
            )

            for text in leg.get_texts():
                plt.setp(text, color=artist_dict["textcolor"])

    def draw_horizontal_bars(self, x, y, **opts_dict):
        """Draw horizontal bar chart with optional error bars and labels."""
        artist_dict = self.plot_dict.copy()

        if isinstance(x, (int, float)):
            x = [x]
        if isinstance(y, (int, float)):
            y = [y]

        for key in opts_dict:
            if key not in artist_dict:
                raise ValueError(key, "is not a valid plotting parameter.")
            artist_dict[key] = opts_dict[key]

        if artist_dict["ignore_sheet_data"] is False:
            if artist_dict["sheet_label"] is not None:
                self.plot_data["label"].append(artist_dict["sheet_label"])
            else:
                self.plot_data["label"].append(artist_dict["label"])

            self.plot_data["x"].append(x)
            self.plot_data["y"].append(y)

        if type(artist_dict["lw"]) is not list:
            artist_dict["lw"] = [artist_dict["lw"]] * len(x)

        if type(artist_dict["lc"]) is not list:
            artist_dict["lc"] = [artist_dict["lc"]] * len(x)

        if type(artist_dict["alpha"]) is not list:
            artist_dict["alpha"] = [artist_dict["alpha"]] * len(x)

        if type(artist_dict['horizontal_bar_left']) is not list:
            artist_dict['horizontal_bar_left'] = (
                [artist_dict['horizontal_bar_left']] * len(x)
            )

        for i in range(len(x)):
            self.current_zorder += 1
            bar_label = (
                artist_dict["label"]
                if i == 0 and artist_dict["label"] is not None
                else None
            )
            self.ax.barh(
                y[i], x[i],
                left=artist_dict['horizontal_bar_left'][i],
                edgecolor=None, lw=None,
                facecolor=artist_dict["lc"][i],
                alpha=artist_dict["alpha"][i],
                align='center', rasterized=True,
                height=artist_dict["horizontal_bar_height"],
                label=bar_label,
                zorder=self.current_zorder,
            )

            if artist_dict["xerr"] is not None:
                self.current_zorder += 1
                self.ax.errorbar(
                    x[i], y[i],
                    xerr=artist_dict["xerr"][i],
                    elinewidth=artist_dict["elw"],
                    ecolor=artist_dict["lc"][i],
                    fmt='none',
                    capsize=artist_dict["elw"] * 1.5,
                    rasterized=artist_dict["rasterized"],
                    mew=artist_dict["elw"],
                    solid_capstyle='round',
                    solid_joinstyle='round',
                    zorder=self.current_zorder,
                )

        if artist_dict["label"] is not None:
            leg = self.ax.legend(
                frameon=False, loc='upper left',
                bbox_to_anchor=(
                    artist_dict["legend_xpos"]
                    / artist_dict["fig_width"],
                    artist_dict["legend_ypos"]
                    / artist_dict["fig_height"],
                ),
                bbox_transform=self.figure.fig.transFigure,
            )

            for text in leg.get_texts():
                plt.setp(text, color=artist_dict["textcolor"])

    def draw_swarmplot(self, ys, **opts_dict):
        """Draw a seaborn swarm plot from a list of value arrays."""
        artist_dict = self.plot_dict.copy()

        for key in opts_dict:
            if key not in artist_dict:
                raise ValueError(key, "is not a valid plotting parameter.")
            artist_dict[key] = opts_dict[key]

        if artist_dict["ignore_sheet_data"] is False:

            if isinstance(artist_dict["sheet_label"], list):
                sheet_labels = artist_dict["sheet_label"]
            else:
                sheet_labels = [artist_dict["sheet_label"]]*len(ys)
            for i, y in enumerate(ys):

                self.plot_data["label"].append(sheet_labels[i])
                self.plot_data["x"].append(i*np.ones_like(y))
                self.plot_data["y"].append(y)

        if type(artist_dict["pc"]) is not list:
            artist_dict["pc"] = (
                [artist_dict["pc"]]
                * len(artist_dict["xticklabels"])
            )

        lc = sns.color_palette(artist_dict["pc"])

        ys = [pd.Series(y) for y in ys]

        df = pd.concat(
            ys, axis=1,
            keys=range(len(artist_dict["xticks"])),
        ).stack(0)
        df = df.reset_index(level=1)
        df.columns = ["X", "Y"]

        self.current_zorder += 1
        g = sns.swarmplot(
            x="X", y="Y", data=df,
            ax=self.ax, palette=lc, hue="X",
            edgecolor=artist_dict["ec"],
            linewidth=artist_dict["elw"],
            rasterized=artist_dict["rasterized"],
            marker=artist_dict["pt"],
            size=artist_dict["ps"],
            alpha=artist_dict["alpha"],
            zorder=self.current_zorder,
        )
        g.get_legend().remove()

        # As seanborn replaces the labels, set them again
        self.ax.set_xticklabels(self.plot_dict["xticklabels"])
        self.ax.set_yticklabels(self.plot_dict["yticklabels"])
        self.ax.set_xlabel(self.plot_dict["xl"])
        self.ax.set_ylabel(self.plot_dict["yl"])

        self.set_axes_properties()

    def draw_violinplot(self, ys, **opts_dict):
        """Draw a seaborn violin plot from a list of value arrays."""
        artist_dict = self.plot_dict.copy()

        for key in opts_dict:
            if key not in artist_dict:
                raise ValueError(key, "is not a valid plotting parameter.")
            artist_dict[key] = opts_dict[key]

        if artist_dict["ignore_sheet_data"] is False:

            if isinstance(artist_dict["sheet_label"], list):
                sheet_labels = artist_dict["sheet_label"]
            else:
                sheet_labels = [artist_dict["sheet_label"]]*len(ys)
            for i, y in enumerate(ys):

                self.plot_data["label"].append(sheet_labels[i])
                self.plot_data["x"].append(i*np.ones_like(y))
                self.plot_data["y"].append(y)

        if type(artist_dict["alpha"]) is not list:
            artist_dict["alpha"] = (
                [artist_dict["alpha"]]
                * len(artist_dict["xticks"])
            )

        if type(artist_dict["pc"]) is not list:
            artist_dict["pc"] = (
                [artist_dict["pc"]]
                * len(artist_dict["xticks"])
            )

        palette = sns.color_palette(artist_dict["pc"])

        ys_ = [pd.Series(y) for y in ys]

        df = pd.concat(ys_, axis=1, keys=range(len(artist_dict["xticks"]))).stack(0)
        df = df.reset_index(level=1)
        df.columns = ["X", "Y"]

        self.current_zorder += 1
        sns.violinplot(
            x="X", y="Y", data=df, ax=self.ax,
            palette=palette, hue="X", legend=False,
            edgecolor=artist_dict["ec"],
            linewidth=artist_dict["elw"], inner=None,
            rasterized=artist_dict["rasterized"],
            zorder=self.current_zorder,
        )
        #g.get_legend().remove()
        plt.setp(self.ax.collections[-len(ys):], alpha=artist_dict["alpha"])

        # As seanborn replaces the labels, set them again
        self.ax.set_xticklabels(self.plot_dict["xticklabels"])
        self.ax.set_yticklabels(self.plot_dict["yticklabels"])
        self.ax.set_xlabel(self.plot_dict["xl"])
        self.ax.set_ylabel(self.plot_dict["yl"])

        self.set_axes_properties()

    def draw_boxplot(self, ys, **opts_dict):
        """Draw a seaborn box plot from a list of value arrays."""
        artist_dict = self.plot_dict.copy()

        for key in opts_dict:
            if key not in artist_dict:
                raise ValueError(key, "is not a valid plotting parameter.")
            artist_dict[key] = opts_dict[key]

        if type(artist_dict["pc"]) is not list:
            artist_dict["pc"] = [artist_dict["pc"]] * len(artist_dict["xticks"])

        palette = sns.color_palette(artist_dict["pc"])

        ys_ = [pd.Series(y) for y in ys]

        df = pd.concat(ys_, axis=1, keys=range(len(artist_dict["xticks"]))).stack(0)
        df = df.reset_index(level=1)
        df.columns = ["X", "Y"]

        self.current_zorder += 1
        flierprops = {"marker": artist_dict["pt"],
                      "markersize": artist_dict["ps"],
                      "zorder": self.current_zorder,
                      "markerfacecolor": artist_dict["textcolor"],
                      "linestyle": 'none',
                      "markeredgecolor": artist_dict["textcolor"]}

        sns.boxplot(x="X", y="Y", data=df, ax=self.ax, palette=palette, hue="X", legend=False,
                    boxprops={"zorder": self.current_zorder, "alpha": artist_dict["alpha"]},
                    linewidth=artist_dict["elw"],
                    width=artist_dict["vertical_bar_width"],
                    flierprops=flierprops,
                    zorder=self.current_zorder)
        #g.get_legend().remove()

        # As seaborn replaces the labels, set them again
        self.ax.set_xticklabels(self.plot_dict["xticklabels"])
        self.ax.set_yticklabels(self.plot_dict["yticklabels"])
        self.ax.set_xlabel(self.plot_dict["xl"])
        self.ax.set_ylabel(self.plot_dict["yl"])

        self.set_axes_properties()

    def draw_image(self, img, extent, **opts_dict):
        """Display an image array on the axes with the given spatial extent.

        Returns
        -------
            The matplotlib AxesImage object.
        """
        artist_dict = self.plot_dict.copy()

        for key in opts_dict:
            if key not in artist_dict:
                raise ValueError(key, "is not a valid plotting parameter.")
            artist_dict[key] = opts_dict[key]

        self.current_zorder += 1
        if artist_dict["norm_colormap"] is None:
            im = self.ax.imshow(
                np.array(img), extent=extent,
                interpolation=artist_dict["image_interpolation"],
                origin=artist_dict["image_origin"],
                aspect='auto',
                cmap=plt.get_cmap(artist_dict["colormap"]),
                vmin=artist_dict["zmin"],
                vmax=artist_dict["zmax"],
                alpha=artist_dict["alpha"],
                rasterized=artist_dict["rasterized"],
                zorder=self.current_zorder,
            )
        else:
            im = self.ax.imshow(
                np.array(img), extent=extent,
                interpolation=artist_dict["image_interpolation"],
                origin=artist_dict["image_origin"],
                aspect='auto',
                cmap=plt.get_cmap(artist_dict["colormap"]),
                norm=artist_dict["norm_colormap"],
                alpha=artist_dict["alpha"],
                rasterized=artist_dict["rasterized"],
                zorder=self.current_zorder,
            )
        return im

    def draw_pcolormesh(self, x, y, Z, aa=False, shading="gouraud", **opts_dict):
        """Draw a pcolormesh plot on the current axes.

        Parameters
        ----------
            x, y: array-like
                Arrays specifying the grid points for the x and y axes.

            Z: array-like
                2D array of data values to be plotted as a color mesh.

            aa: bool, optional
                Whether to use antialiased rendering. Default is False.

            shading: str, optional
                The shading method to be used. Default is "gouraud".

            **opts_dict (keyword arguments): Additional options for the pcolormesh plot.
                The following options are available:

                - norm_colormap: object or None, optional
                    Normalization object for the colormap. If None,
                    the colormap will not be normalized. Default is None.

                - colormap: str, optional
                    Name of the colormap to use for color mappings. Default is "inferno".

                - rasterized: bool, optional
                    If True, display the plot as a pixel raster for
                    improved rendering and loading of huge data
                    clouds. Default is False.

                - zmin, zmax: float or None, optional
                    Minimum and maximum values for the color scale. Default is None (auto-scaling).

                - alpha: float, optional
                    Transparency of the color mesh. Default is 1.0.

        Returns
        -------
            None
        """
        artist_dict = self.plot_dict.copy()

        for key in opts_dict:
            if key not in artist_dict:
                raise ValueError(key, "is not a valid plotting parameter.")
            artist_dict[key] = opts_dict[key]

        self.current_zorder += 1
        if artist_dict["norm_colormap"] is None:
            self.ax.pcolormesh(x, y, Z, aa=aa, shading=shading,
                               cmap=plt.get_cmap(artist_dict["colormap"]),
                               rasterized=artist_dict["rasterized"],
                               vmin=artist_dict["zmin"],
                               vmax=artist_dict["zmax"],
                               alpha=artist_dict["alpha"],
                               zorder=self.current_zorder)
        else:
            self.ax.pcolormesh(x, y, Z, aa=aa, shading=shading,
                               cmap=plt.get_cmap(artist_dict["colormap"]),
                               rasterized=artist_dict["rasterized"],
                               norm=artist_dict["norm_colormap"],
                               alpha=artist_dict["alpha"],
                               zorder=self.current_zorder)


class PolarPlot(SuperPlot):
    """Polar-coordinate subplot panel."""

    def __init__(self, figure, opts_dict):
        """Initialize polar axes with ticks, grid, and labels."""
        SuperPlot.__init__(self, figure, opts_dict)

        self.current_zorder = 0

        self.ax = self.figure.fig.add_axes(
            [
                self.plot_dict["xpos"] / self.plot_dict["fig_width"],
                self.plot_dict["ypos"] / self.plot_dict["fig_height"],
                (self.plot_dict["plot_width"]
                 / self.plot_dict["fig_width"]),
                (self.plot_dict["plot_height"]
                 / self.plot_dict["fig_height"]),
            ],
            polar=True,
        )

        self.ax.set_facecolor("none")
        self.ax.spines['polar'].set_linewidth(
            self.plot_dict["axes_lw"],
        )
        self.ax.spines['polar'].set_color(
            self.plot_dict["textcolor"],
        )
        self.ax.spines['polar'].set_zorder(100)

        self.ax.set_theta_zero_location('N')
        self.ax.set_theta_direction(-1)

        self.current_zorder += 1
        self.ax.grid(
            color=self.plot_dict["helper_lines_lc"],
            linewidth=self.plot_dict["axes_lw"] * 0.5,
            dashes=self.plot_dict["helper_lines_dashes"],
            solid_capstyle="round",
            dash_capstyle=self.plot_dict["dash_capstyle"],
            zorder=self.current_zorder,
        )

        self.ax.set_rticks(self.plot_dict["yticks"])# zorder=100)
        self.ax.set_rlim([self.plot_dict["ymin"], self.plot_dict["ymax"]])

        if self.plot_dict["xticklabels"] is None:
            self.plot_dict["xticklabels"] = [str(lbl) for lbl in self.plot_dict["xticks"]]

        if self.plot_dict["yticklabels"] is None:
            self.plot_dict["yticklabels"] = [str(lbl) for lbl in self.plot_dict["yticks"]]

        for i in range(len(self.plot_dict["xticklabels"])):
            self.plot_dict["xticklabels"][i] = self.plot_dict["xticklabels"][i].replace("-", '–')

        for i in range(len(self.plot_dict["yticklabels"])):
            self.plot_dict["yticklabels"][i] = self.plot_dict["yticklabels"][i].replace("-", '–')

        self.ax.set_xlim([self.plot_dict["xmin"], self.plot_dict["xmax"]])
        self.ax.set_xticks(self.plot_dict["xticks"])#zorder=100)

        if self.plot_dict["xticklabels_rotation"] == 0:
            self.ax.set_xticklabels(
                self.plot_dict["xticklabels"],
                rotation=self.plot_dict["xticklabels_rotation"],
                horizontalalignment='center',
                color=self.plot_dict["textcolor"],
            )
        else:
            self.ax.set_xticklabels(
                self.plot_dict["xticklabels"],
                rotation=self.plot_dict["xticklabels_rotation"],
                horizontalalignment='right',
                color=self.plot_dict["textcolor"],
            )

        if self.plot_dict["yticklabels_rotation"] == 0:
            self.ax.set_yticklabels(
                self.plot_dict["yticklabels"],
                rotation=self.plot_dict["yticklabels_rotation"],
                horizontalalignment='center',
                color=self.plot_dict["textcolor"],
            )
        else:
            self.ax.set_yticklabels(
                self.plot_dict["yticklabels"],
                rotation=self.plot_dict["yticklabels_rotation"],
                horizontalalignment='right',
                color=self.plot_dict["textcolor"],
            )

        self.current_zorder += 1
        tick = [self.ax.get_rmax(), self.ax.get_rmax() * 0.97]
        for t in self.plot_dict["xticks"]:
            self.ax.plot(
                [t, t], tick,
                lw=self.plot_dict["axes_lw"],
                color=self.plot_dict["textcolor"],
                zorder=self.current_zorder,
            )

        plt.figtext(
            (self.plot_dict["xpos"]
             - 0.3 * self.plot_dict["fontsize"] / 9.)
            / self.plot_dict["fig_width"],
            (self.plot_dict["ypos"]
             + self.plot_dict["plot_height"] + 0.5)
            / self.plot_dict["fig_height"],
            self.plot_dict["plot_label"],
            weight='bold',
            fontsize=self.plot_dict["fontsize_plot_label"],
            ha='center', va='center',
            color=self.plot_dict["textcolor"],
        )

        self.set_axes_properties()

class Plot(SuperPlot):
    """Standard 2D cartesian subplot panel."""

    def __init__(self, figure, opts_dict):
        """Initialize 2D axes with spines, ticks, helper lines, and labels."""
        SuperPlot.__init__(self, figure, opts_dict)

        self.ax = self.figure.fig.add_axes([
            self.plot_dict["xpos"] / self.plot_dict["fig_width"],
            self.plot_dict["ypos"] / self.plot_dict["fig_height"],
            (self.plot_dict["plot_width"]
             / self.plot_dict["fig_width"]),
            (self.plot_dict["plot_height"]
             / self.plot_dict["fig_height"]),
        ])

        self.ax.set_facecolor("none")

        self.ax.spines['right'].set_color('none')
        self.ax.spines['top'].set_color('none')

        self.ax.spines['left'].set_linewidth(
            self.plot_dict["axes_lw"],
        )
        self.ax.spines['bottom'].set_linewidth(
            self.plot_dict["axes_lw"],
        )
        self.ax.spines['left'].set_color(
            self.plot_dict["textcolor"],
        )
        self.ax.spines['bottom'].set_color(
            self.plot_dict["textcolor"],
        )
        self.ax.spines["left"].set_zorder(100)
        self.ax.spines["bottom"].set_zorder(100)

        self.ax.xaxis.set_ticks_position('bottom')
        self.ax.yaxis.set_ticks_position('left')
        self.ax.tick_params(
            'both', width=self.plot_dict["axes_lw"],
            which='major', tickdir="out",
            color=self.plot_dict["textcolor"],
            zorder=100,
        )
        self.ax.tick_params(
            'both', width=self.plot_dict["axes_lw"],
            which='minor', tickdir="out",
            color=self.plot_dict["textcolor"],
            zorder=100,
        )

        if self.plot_dict["xticks"] is not None:
            self.ax.set_xticks(self.plot_dict["xticks"])

            if self.plot_dict["xticklabels"] is None:
                self.plot_dict["xticklabels"] = [
                    str(lbl)
                    for lbl in self.plot_dict["xticks"]
                ]

        if self.plot_dict["xticklabels"] is not None:
            self.plot_dict["xticklabels"] = (
                self.plot_dict["xticklabels"].copy()
            )

            for i in range(len(self.plot_dict["xticklabels"])):
                self.plot_dict["xticklabels"][i] = (
                    self.plot_dict["xticklabels"][i]
                    .replace("-", '\u2013')
                )

            if self.plot_dict["xticklabels_rotation"] == 0:
                self.ax.set_xticklabels(
                    self.plot_dict["xticklabels"],
                    rotation=(
                        self.plot_dict["xticklabels_rotation"]
                    ),
                    horizontalalignment='center',
                    color=self.plot_dict["textcolor"],
                )
            else:
                self.ax.set_xticklabels(
                    self.plot_dict["xticklabels"],
                    rotation=(
                        self.plot_dict["xticklabels_rotation"]
                    ),
                    horizontalalignment='right',
                    color=self.plot_dict["textcolor"],
                )

        else:
            self.ax.spines['bottom'].set_visible(False)
            self.ax.tick_params(axis='x', which='minor', bottom='off')
            self.ax.tick_params(axis='x', which='major', bottom='off')
            self.ax.get_xaxis().set_ticks([])

        if self.plot_dict["yticks"] is not None:
            self.ax.set_yticks(self.plot_dict["yticks"])

            if self.plot_dict["yticklabels"] is None:
                self.plot_dict["yticklabels"] = [
                    str(lbl)
                    for lbl in self.plot_dict["yticks"]
                ]

        if self.plot_dict["yticklabels"] is not None:
            self.plot_dict["yticklabels"] = (
                self.plot_dict["yticklabels"].copy()
            )

            for i in range(len(self.plot_dict["yticklabels"])):
                self.plot_dict["yticklabels"][i] = (
                    self.plot_dict["yticklabels"][i]
                    .replace("-", '\u2013')
                )

            self.ax.set_yticklabels(
                self.plot_dict["yticklabels"],
                rotation=(
                    self.plot_dict["yticklabels_rotation"]
                ),
                horizontalalignment='right',
                color=self.plot_dict["textcolor"],
            )
        else:
            self.ax.spines['left'].set_visible(False)
            self.ax.tick_params(
                axis='y', which='minor', left='off',
            )
            self.ax.tick_params(
                axis='y', which='major', left='off',
            )
            self.ax.get_yaxis().set_ticks([])

        if self.plot_dict["vspans"] is not None:
            self.current_zorder += 1
            for vspan in self.plot_dict["vspans"]:
                self.ax.axvspan(
                    vspan[0], vspan[1], lw=0,
                    edgecolor='none',
                    facecolor=vspan[2],
                    alpha=vspan[3],
                    zorder=self.current_zorder,
                )

        if self.plot_dict["hspans"] is not None:
            self.current_zorder += 1
            for hspan in self.plot_dict["hspans"]:
                self.ax.axhspan(
                    hspan[0], hspan[1], lw=0,
                    edgecolor='none',
                    facecolor=hspan[2],
                    alpha=hspan[3],
                    zorder=self.current_zorder,
                )

        if self.plot_dict["hlines"] is not None:
            self.current_zorder += 1
            for hline in self.plot_dict["hlines"]:
                self.ax.axhline(
                    hline,
                    linewidth=self.plot_dict["helper_lines_lw"],
                    color=self.plot_dict["helper_lines_lc"],
                    dashes=self.plot_dict["helper_lines_dashes"],
                    alpha=self.plot_dict["helper_lines_alpha"],
                    solid_capstyle="round",
                    dash_capstyle=(
                        self.plot_dict["dash_capstyle"]
                    ),
                    zorder=self.current_zorder,
                )

        if self.plot_dict["vlines"] is not None:
            self.current_zorder += 1
            for vline in self.plot_dict["vlines"]:
                self.ax.axvline(
                    vline,
                    linewidth=self.plot_dict["helper_lines_lw"],
                    color=self.plot_dict["helper_lines_lc"],
                    dashes=self.plot_dict["helper_lines_dashes"],
                    alpha=self.plot_dict["helper_lines_alpha"],
                    solid_capstyle="round",
                    dash_capstyle=(
                        self.plot_dict["dash_capstyle"]
                    ),
                    zorder=self.current_zorder,
                )

        if self.plot_dict["plot_title"] is not None:
            self.ax.set_title(
                self.plot_dict["plot_title"],
                color=self.plot_dict["textcolor"],
                fontsize=self.plot_dict["fontsize"],
            )

        if self.plot_dict["plot_label"] is not None:
            if self.ax.spines['left'].get_visible():
                plt.figtext(
                    (self.plot_dict["xpos"]
                     - 1.8 * self.plot_dict["fontsize"] / 9.)
                    / self.plot_dict["fig_width"],
                    (self.plot_dict["ypos"]
                     + self.plot_dict["plot_height"] + 0.5)
                    / self.plot_dict["fig_height"],
                    self.plot_dict["plot_label"],
                    weight='bold',
                    fontsize=(
                        self.plot_dict["fontsize_plot_label"]
                    ),
                    ha='center', va='center',
                    color=self.plot_dict["textcolor"],
                )
            else:
                plt.figtext(
                    (self.plot_dict["xpos"]
                     - 0.3 * self.plot_dict["fontsize"] / 9.)
                    / self.plot_dict["fig_width"],
                    (self.plot_dict["ypos"]
                     + self.plot_dict["plot_height"] + 0.5)
                    / self.plot_dict["fig_height"],
                    self.plot_dict["plot_label"],
                    weight='bold',
                    fontsize=(
                        self.plot_dict["fontsize_plot_label"]
                    ),
                    ha='center', va='center',
                    color=self.plot_dict["textcolor"],
                )

        # Draw the colormap next to it
        if self.plot_dict["show_colormap"]:
            cbar_ax = self.figure.fig.add_axes(
                [
                    (self.plot_dict["xpos"]
                     + self.plot_dict["plot_width"]
                     + self.plot_dict["plot_width"] / 20.)
                    / self.plot_dict["fig_width"],
                    (self.plot_dict["ypos"]
                     / self.plot_dict["fig_height"]),
                    (self.plot_dict["plot_width"] / 10.)
                    / self.plot_dict["fig_width"],
                    (self.plot_dict["plot_height"]
                     / self.plot_dict["fig_height"]),
                ],
                frameon=False,
            )

            cbar_ax.yaxis.set_ticks([])
            cbar_ax.xaxis.set_ticks([])

            cbar_ax2 = cbar_ax.twinx()
            cbar_ax2.set_facecolor("none")

            for val in ["left", "right", "bottom", "top"]:
                cbar_ax2.spines[val].set_zorder(100)
                cbar_ax2.spines[val].set_linewidth(
                    self.plot_dict["axes_lw"],
                )
                cbar_ax2.spines[val].set_color(
                    self.plot_dict["textcolor"],
                )

            cbar_ax2.tick_params(
                'both',
                width=self.plot_dict["axes_lw"],
                which='major', tickdir="out",
                color=self.plot_dict["textcolor"],
                zorder=100,
            )
            cbar_ax2.tick_params(
                'both',
                width=self.plot_dict["axes_lw"],
                which='minor', tickdir="out",
                color=self.plot_dict["textcolor"],
                zorder=100,
            )

            self.current_zorder += 1
            if self.plot_dict["norm_colormap"] is None:
                cbar_ax2.imshow(
                    np.c_[np.linspace(
                        self.plot_dict["zmin"],
                        self.plot_dict["zmax"], 500,
                    )],
                    extent=(
                        0, 1,
                        self.plot_dict["zmin"],
                        self.plot_dict["zmax"],
                    ),
                    alpha=self.plot_dict["alpha"],
                    rasterized=self.plot_dict["rasterized"],
                    aspect='auto', origin='lower',
                    cmap=plt.get_cmap(
                        self.plot_dict["colormap"],
                    ),
                    zorder=self.current_zorder,
                )
            else:
                norm_cm = self.plot_dict["norm_colormap"]
                cbar_ax2.imshow(
                    np.c_[np.linspace(
                        norm_cm.vmin, norm_cm.vmax, 500,
                    )],
                    extent=(
                        0, 1, norm_cm.vmin, norm_cm.vmax,
                    ),
                    alpha=self.plot_dict["alpha"],
                    rasterized=self.plot_dict["rasterized"],
                    aspect='auto', origin='lower',
                    norm=norm_cm,
                    cmap=plt.get_cmap(
                        self.plot_dict["colormap"],
                    ),
                    zorder=self.current_zorder,
                )

            if self.plot_dict["zticks"] is not None:
                cbar_ax2.set_yticks(
                    self.plot_dict["zticks"],
                )

                if self.plot_dict["zticklabels"] is None:
                    self.plot_dict["zticklabels"] = [
                        str(lbl)
                        for lbl in self.plot_dict["zticks"]
                    ]

                for i in range(
                    len(self.plot_dict["zticklabels"])
                ):
                    self.plot_dict["zticklabels"][i] = (
                        self.plot_dict["zticklabels"][i]
                        .replace("-", '\u2013')
                    )

                cbar_ax2.set_yticklabels(
                    self.plot_dict["zticklabels"],
                    rotation=(
                        self.plot_dict["zticklabels_rotation"]
                    ),
                    horizontalalignment=(
                        self.plot_dict["zticklabels_ha"]
                    ),
                    color=self.plot_dict["textcolor"],
                )

            if self.plot_dict["zl"] is not None:
                cbar_ax2.set_ylabel(
                    self.plot_dict["zl"],
                    color=self.plot_dict["textcolor"],
                )

        self.set_axes_properties()


class Plot3D(SuperPlot):
    """3D cartesian subplot panel using matplotlib Axes3D."""

    def __init__(self, figure, opts_dict):
        """Initialize 3D axes with spines, grid, ticks, and view angle."""
        SuperPlot.__init__(self, figure, opts_dict)

        self.ax = self.figure.fig.add_axes(
            [
                self.plot_dict["xpos"]
                / self.plot_dict["fig_width"],
                self.plot_dict["ypos"]
                / self.plot_dict["fig_height"],
                (self.plot_dict["plot_width"]
                 / self.plot_dict["fig_width"]),
                (self.plot_dict["plot_height"]
                 / self.plot_dict["fig_height"]),
            ],
            projection='3d',
        )

        self.ax.set_facecolor("none")

        self.ax.spines['right'].set_color('none')
        self.ax.spines['top'].set_color('none')

        self.ax.spines['left'].set_linewidth(
            self.plot_dict["axes_lw"],
        )
        self.ax.spines['bottom'].set_linewidth(
            self.plot_dict["axes_lw"],
        )
        self.ax.spines['left'].set_color(
            self.plot_dict["textcolor"],
        )
        self.ax.spines['bottom'].set_color(
            self.plot_dict["textcolor"],
        )
        self.ax.spines["left"].set_zorder(100)
        self.ax.spines["bottom"].set_zorder(100)

        self.ax.xaxis.set_ticks_position('lower')
        self.ax.yaxis.set_ticks_position('lower')
        self.ax.zaxis.set_ticks_position('lower')
        self.ax.grid(
            color=self.plot_dict["helper_lines_lc"],
            linewidth=self.plot_dict["axes_lw"] * 0.5,
            dashes=self.plot_dict["helper_lines_dashes"],
            solid_capstyle="round",
            dash_capstyle=(
                self.plot_dict["dash_capstyle"]
            ),
            zorder=self.current_zorder,
        )
        self.ax.tick_params(
            'both', width=self.plot_dict["axes_lw"],
            which='major', tickdir="out",
            color=self.plot_dict["textcolor"],
            zorder=100,
        )
        self.ax.tick_params(
            'both', width=self.plot_dict["axes_lw"],
            which='minor', tickdir="out",
            color=self.plot_dict["textcolor"],
            zorder=100,
        )

        if (
            self.plot_dict["elev"] is not None
            and self.plot_dict["azim"] is not None
        ):
            self.ax.view_init(
                elev=self.plot_dict["elev"],
                azim=self.plot_dict["azim"],
            )
        elif self.plot_dict["elev"] is not None:
            self.ax.view_init(
                elev=self.plot_dict["elev"],
            )
        elif self.plot_dict["azim"] is not None:
            self.ax.view_init(
                azim=self.plot_dict["azim"],
            )

        if self.plot_dict["xticks"] is not None:
            self.ax.set_xticks(self.plot_dict["xticks"])

            if self.plot_dict["xticklabels"] is None:
                self.plot_dict["xticklabels"] = [
                    str(lbl)
                    for lbl in self.plot_dict["xticks"]
                ]

        if self.plot_dict["xticklabels"] is not None:
            self.plot_dict["xticklabels"] = (
                self.plot_dict["xticklabels"].copy()
            )

            for i in range(len(self.plot_dict["xticklabels"])):
                self.plot_dict["xticklabels"][i] = (
                    self.plot_dict["xticklabels"][i]
                    .replace("-", '\u2013')
                )

            if self.plot_dict["xticklabels_rotation"] == 0:
                self.ax.set_xticklabels(
                    self.plot_dict["xticklabels"],
                    rotation=(
                        self.plot_dict["xticklabels_rotation"]
                    ),
                    horizontalalignment='center',
                    color=self.plot_dict["textcolor"],
                )
            else:
                self.ax.set_xticklabels(
                    self.plot_dict["xticklabels"],
                    rotation=(
                        self.plot_dict["xticklabels_rotation"]
                    ),
                    horizontalalignment='right',
                    color=self.plot_dict["textcolor"],
                )

        else:
            self.ax.spines['bottom'].set_visible(False)
            self.ax.tick_params(
                axis='x', which='minor', bottom='off',
            )
            self.ax.tick_params(
                axis='x', which='major', bottom='off',
            )
            self.ax.get_xaxis().set_ticks([])

        if self.plot_dict["yticks"] is not None:
            self.ax.set_yticks(self.plot_dict["yticks"])

            if self.plot_dict["yticklabels"] is None:
                self.plot_dict["yticklabels"] = [
                    str(lbl)
                    for lbl in self.plot_dict["yticks"]
                ]

        if self.plot_dict["yticklabels"] is not None:
            self.plot_dict["yticklabels"] = (
                self.plot_dict["yticklabels"].copy()
            )

            for i in range(len(self.plot_dict["yticklabels"])):
                self.plot_dict["yticklabels"][i] = (
                    self.plot_dict["yticklabels"][i]
                    .replace("-", '\u2013')
                )

            self.ax.set_yticklabels(
                self.plot_dict["yticklabels"],
                rotation=(
                    self.plot_dict["yticklabels_rotation"]
                ),
                horizontalalignment='right',
                color=self.plot_dict["textcolor"],
            )
        else:
            self.ax.spines['left'].set_visible(False)
            self.ax.tick_params(
                axis='y', which='minor', left='off',
            )
            self.ax.tick_params(
                axis='y', which='major', left='off',
            )
            self.ax.get_yaxis().set_ticks([])

        if self.plot_dict["zticks"] is not None:
            self.ax.set_zticks(self.plot_dict["zticks"])

            if self.plot_dict["zticklabels"] is None:
                self.plot_dict["zticklabels"] = [
                    str(lbl)
                    for lbl in self.plot_dict["zticks"]
                ]

        if self.plot_dict["zticklabels"] is not None:
            self.plot_dict["zticklabels"] = (
                self.plot_dict["zticklabels"].copy()
            )

            for i in range(len(self.plot_dict["zticklabels"])):
                self.plot_dict["zticklabels"][i] = (
                    self.plot_dict["zticklabels"][i]
                    .replace("-", '\u2013')
                )

            self.ax.set_zticklabels(
                self.plot_dict["zticklabels"],
                rotation=(
                    self.plot_dict["zticklabels_rotation"]
                ),
                horizontalalignment='right',
                color=self.plot_dict["textcolor"],
            )
        else:
            self.ax.spines['left'].set_visible(False)
            self.ax.tick_params(
                axis='z', which='minor', left='off',
            )
            self.ax.tick_params(
                axis='z', which='major', left='off',
            )
            self.ax.get_zaxis().set_ticks([])

        if self.plot_dict["vspans"] is not None:
            self.current_zorder += 1
            for vspan in self.plot_dict["vspans"]:
                self.ax.axvspan(
                    vspan[0], vspan[1], lw=0,
                    edgecolor='none',
                    facecolor=vspan[2],
                    alpha=vspan[3],
                    zorder=self.current_zorder,
                )

        if self.plot_dict["hspans"] is not None:
            self.current_zorder += 1
            for hspan in self.plot_dict["hspans"]:
                self.ax.axhspan(
                    hspan[0], hspan[1], lw=0,
                    edgecolor='none',
                    facecolor=hspan[2],
                    alpha=hspan[3],
                    zorder=self.current_zorder,
                )

        if self.plot_dict["hlines"] is not None:
            self.current_zorder += 1
            for hline in self.plot_dict["hlines"]:
                self.ax.axhline(
                    hline,
                    linewidth=self.plot_dict["helper_lines_lw"],
                    color=self.plot_dict["helper_lines_lc"],
                    dashes=self.plot_dict["helper_lines_dashes"],
                    alpha=self.plot_dict["helper_lines_alpha"],
                    solid_capstyle="round",
                    dash_capstyle=(
                        self.plot_dict["dash_capstyle"]
                    ),
                    zorder=self.current_zorder,
                )

        if self.plot_dict["vlines"] is not None:
            self.current_zorder += 1
            for vline in self.plot_dict["vlines"]:
                self.ax.axvline(
                    vline,
                    linewidth=self.plot_dict["helper_lines_lw"],
                    color=self.plot_dict["helper_lines_lc"],
                    dashes=self.plot_dict["helper_lines_dashes"],
                    alpha=self.plot_dict["helper_lines_alpha"],
                    solid_capstyle="round",
                    dash_capstyle=(
                        self.plot_dict["dash_capstyle"]
                    ),
                    zorder=self.current_zorder,
                )

        if self.plot_dict["plot_title"] is not None:
            self.ax.set_title(
                self.plot_dict["plot_title"],
                color=self.plot_dict["textcolor"],
                fontsize=self.plot_dict["fontsize"],
            )

        if self.plot_dict["plot_label"] is not None:
            if self.ax.spines['left'].get_visible():
                plt.figtext(
                    (self.plot_dict["xpos"]
                     - 1.8 * self.plot_dict["fontsize"] / 9.)
                    / self.plot_dict["fig_width"],
                    (self.plot_dict["ypos"]
                     + self.plot_dict["plot_height"] + 0.5)
                    / self.plot_dict["fig_height"],
                    self.plot_dict["plot_label"],
                    weight='bold',
                    fontsize=(
                        self.plot_dict["fontsize_plot_label"]
                    ),
                    ha='center', va='center',
                    color=self.plot_dict["textcolor"],
                )
            else:
                plt.figtext(
                    (self.plot_dict["xpos"]
                     - 0.3 * self.plot_dict["fontsize"] / 9.)
                    / self.plot_dict["fig_width"],
                    (self.plot_dict["ypos"]
                     + self.plot_dict["plot_height"] + 0.5)
                    / self.plot_dict["fig_height"],
                    self.plot_dict["plot_label"],
                    weight='bold',
                    fontsize=(
                        self.plot_dict["fontsize_plot_label"]
                    ),
                    ha='center', va='center',
                    color=self.plot_dict["textcolor"],
                )

        # Draw the colormap next to it
        if self.plot_dict["show_colormap"]:
            cbar_ax = self.figure.fig.add_axes(
                [
                    (self.plot_dict["xpos"]
                     + self.plot_dict["plot_width"]
                     + self.plot_dict["plot_width"] / 20.)
                    / self.plot_dict["fig_width"],
                    (self.plot_dict["ypos"]
                     / self.plot_dict["fig_height"]),
                    (self.plot_dict["plot_width"] / 10.)
                    / self.plot_dict["fig_width"],
                    (self.plot_dict["plot_height"]
                     / self.plot_dict["fig_height"]),
                ],
                frameon=False,
            )

            cbar_ax.yaxis.set_ticks([])
            cbar_ax.xaxis.set_ticks([])
            cbar_ax.zaxis.set_ticks([])

            cbar_ax2 = cbar_ax.twinx()
            cbar_ax2.set_facecolor("none")

            for val in ["left", "right", "bottom", "top"]:
                cbar_ax2.spines[val].set_zorder(100)
                cbar_ax2.spines[val].set_linewidth(
                    self.plot_dict["axes_lw"],
                )
                cbar_ax2.spines[val].set_color(
                    self.plot_dict["textcolor"],
                )

            cbar_ax2.tick_params(
                'both',
                width=self.plot_dict["axes_lw"],
                which='major', tickdir="out",
                color=self.plot_dict["textcolor"],
                zorder=100,
            )
            cbar_ax2.tick_params(
                'both',
                width=self.plot_dict["axes_lw"],
                which='minor', tickdir="out",
                color=self.plot_dict["textcolor"],
                zorder=100,
            )

            self.current_zorder += 1
            if self.plot_dict["norm_colormap"] is None:
                cbar_ax2.imshow(
                    np.c_[np.linspace(
                        self.plot_dict["zmin"],
                        self.plot_dict["zmax"], 500,
                    )],
                    extent=(
                        0, 1,
                        self.plot_dict["zmin"],
                        self.plot_dict["zmax"],
                    ),
                    alpha=self.plot_dict["alpha"],
                    rasterized=self.plot_dict["rasterized"],
                    aspect='auto', origin='lower',
                    cmap=plt.get_cmap(
                        self.plot_dict["colormap"],
                    ),
                    zorder=self.current_zorder,
                )
            else:
                norm_cm = self.plot_dict["norm_colormap"]
                cbar_ax2.imshow(
                    np.c_[np.linspace(
                        norm_cm.vmin, norm_cm.vmax, 500,
                    )],
                    extent=(
                        0, 1, norm_cm.vmin, norm_cm.vmax,
                    ),
                    alpha=self.plot_dict["alpha"],
                    rasterized=self.plot_dict["rasterized"],
                    aspect='auto', origin='lower',
                    norm=norm_cm,
                    cmap=plt.get_cmap(
                        self.plot_dict["colormap"],
                    ),
                    zorder=self.current_zorder,
                )

            if self.plot_dict["zticks"] is not None:
                cbar_ax2.set_yticks(
                    self.plot_dict["zticks"],
                )

                if self.plot_dict["zticklabels"] is None:
                    self.plot_dict["zticklabels"] = [
                        str(lbl)
                        for lbl in self.plot_dict["zticks"]
                    ]

                for i in range(
                    len(self.plot_dict["zticklabels"])
                ):
                    self.plot_dict["zticklabels"][i] = (
                        self.plot_dict["zticklabels"][i]
                        .replace("-", '\u2013')
                    )

                cbar_ax2.set_yticklabels(
                    self.plot_dict["zticklabels"],
                    rotation=(
                        self.plot_dict["zticklabels_rotation"]
                    ),
                    horizontalalignment=(
                        self.plot_dict["zticklabels_ha"]
                    ),
                    color=self.plot_dict["textcolor"],
                )

            if self.plot_dict["zl"] is not None:
                cbar_ax2.set_ylabel(
                    self.plot_dict["zl"],
                    color=self.plot_dict["textcolor"],
                )

        self.set_axes_properties()


